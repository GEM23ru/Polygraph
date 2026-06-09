"""
Polygraph Agent — слой инструментов (function calling).

Поддерживает:
  - OpenRouter (через OpenAI-совместимый tools API)
  - Google Gemini (через нативный function calling)
"""

import json, time, os, inspect
from typing import Any
from dataclasses import dataclass


@dataclass
class Tool:
    name: str
    description: str
    parameters: dict
    func: Any

    def call(self, **kw) -> str:
        try:
            # Отфильтровываем аргументы, которых нет в сигнатуре функции,
            # чтобы лишний/неверно названный ключ от модели не ронял вызов.
            sig = inspect.signature(self.func)
            accepts_kwargs = any(
                p.kind == inspect.Parameter.VAR_KEYWORD
                for p in sig.parameters.values()
            )
            if not accepts_kwargs:
                allowed = set(sig.parameters)
                kw = {k: v for k, v in kw.items() if k in allowed}
            return str(self.func(**kw))
        except Exception as e:
            return f"[tool error] {e}"


class Agent:
    """Инструментальный слой над Polygraph."""

    def __init__(self, pg, debug: bool = False):
        self.pg = pg
        self.tools: dict[str, Tool] = {}
        self.debug = debug
        self.last_model = ""  # какая модель ответила в последний раз
        self.force_model = ""  # принудительная модель (алиас), напр. "gpt-oss" / "gemini"
        self.session_model = None  # (provider, model_id) — закреплённая на сессию модель

    def register(self, tool: Tool):
        self.tools[tool.name] = tool

    def _tools_schema(self) -> list[dict]:
        return [{"type":"function","function":{
            "name":t.name,"description":t.description,"parameters":t.parameters
        }} for t in self.tools.values()]

    # ===== Google Gemini (нативный function calling) =====
    def _run_via_google(self, msg: str, max_turns: int = 5) -> str | None:
        prov = self.pg.providers.get("google")
        if not prov or not prov.ok:
            return None

        # Конвертируем инструменты в формат Gemini
        from google.genai import types as gt
        gemini_tools = []
        for t in self.tools.values():
            # Конвертируем JSON Schema параметры в формат Gemini Schema
            gemini_tools.append(gt.Tool(
                function_declarations=[gt.FunctionDeclaration(
                    name=t.name,
                    description=t.description,
                    parameters=json.loads(json.dumps(t.parameters))  # через JSON чтобы убрать лишнее
                )]
            ))

        config = gt.GenerateContentConfig(
            tools=gemini_tools,
            max_output_tokens=2048,
        )

        contents = [
            self.SYS_PROMPT,
            msg
        ]

        for _ in range(max_turns):
            try:
                resp = prov.c.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=contents,
                    config=config,
                )

                # Проверяем function calls
                if resp.candidates and resp.candidates[0].content and resp.candidates[0].content.parts:
                    parts = resp.candidates[0].content.parts
                    fc_found = False
                    response_text = ""

                    # Сначала добавляем ответ модели (с function_call) в историю,
                    # иначе Gemini не сможет сопоставить function_response.
                    contents.append(resp.candidates[0].content)

                    for part in parts:
                        if hasattr(part, 'function_call') and part.function_call:
                            fc_found = True
                            fname = part.function_call.name
                            fargs = dict(part.function_call.args or {})
                            # Выполняем инструмент РОВНО один раз
                            result = self._execute_tool(fname, fargs)
                            if self.debug:
                                print(f"   [gemini tool] {fname}({fargs}) -> {result[:120]}")
                            contents.append(gt.Part.from_function_response(
                                name=fname,
                                response={"result": result},
                            ))
                        elif hasattr(part, 'text') and part.text:
                            response_text += part.text

                    if not fc_found:
                        self.last_model = "google/gemini-2.5-flash"
                        return response_text or "[no response]"
                else:
                    self.last_model = "google/gemini-2.5-flash"
                    return str(resp.text) if hasattr(resp, 'text') else "[no response]"

            except Exception as e:
                if self.debug:
                    print(f"   [gemini error] {type(e).__name__}: {e}")
                return None  # Пробуем следующий провайдер

        return "[max turns]"

    def _execute_tool(self, name: str, args: dict) -> str:
        tool = self.tools.get(name)
        if tool:
            return tool.call(**args)
        return f"[tool not found: {name}]"

    # ===== OpenAI-совместимый function calling (Groq, OpenRouter, Cerebras, Mistral) =====
    SYS_PROMPT = (
        "Ты — умный, дружелюбный и живой ассистент. Общаешься на «ты», тепло и "
        "по-человечески, как хороший помощник, а не робот-справочник. У тебя есть "
        "инструменты, и ты сам решаешь, когда они нужны. Сейчас 2026 год.\n\n"
        "СТИЛЬ И ПОВЕДЕНИЕ:\n"
        "• Говори естественно, живым языком, лёгкие эмодзи где уместно. 🙂\n"
        "• Здоровайся ТОЛЬКО в первом сообщении беседы. Если в истории уже есть твои "
        "реплики — не здоровайся, сразу к делу.\n"
        "• Длину ответа подстраивай под вопрос: простой/бытовой → коротко, без простыней "
        "и таблиц; развёрнутый разбор (разделы, таблицы, источники) — только когда просят "
        "анализ/сравнение/исследование или это явно нужно.\n"
        "• ⚠️ СНАЧАЛА ответь на сам вопрос, и только ПОТОМ (по желанию) предложи следующий "
        "шаг. Никогда не начинай с «что ещё?» вместо ответа. На «2+2» отвечай «4», на "
        "«15% от 3400» — «510 ₽», и лишь затем, если хочешь, доп. вопрос.\n"
        "• Если запрос размытый/неоднозначный или не хватает деталей — не гадай вслепую, "
        "задай 1-2 коротких уточняющих вопроса (можно с вариантами), потом действуй. "
        "Если ясно — не тяни, помогай сразу.\n\n"
        "🧭 КАК ДУМАТЬ НАД ЗАПРОСОМ (про себя): пойми намерение → реши, нужны ли "
        "инструменты → если задача сложная, составь план из шагов (найти → прочитать → "
        "посчитать → ответить) → выполни, дождись РЕАЛЬНЫХ результатов → синтезируй ответ "
        "только на их основе. Типы: болтовня/приветствие и общеизвестный факт → без "
        "инструментов; актуальное/новости/«найди»/обзор/сравнение → deep_research; "
        "расчёты/анализ/код → run_python; описать картинку → analyze_image; нарисовать → "
        "generate_image; файлы → read_file/write_file/list_files. Не зови инструмент зря, "
        "но для фактов из интернета/расчётов/картинок — зови обязательно.\n\n"
        "ИНСТРУМЕНТЫ:\n"
        "• deep_research — глубокое исследование (поиск + чтение страниц + источники). "
        "web_search — быстрый поиск. fetch_url — прочитать страницу по ссылке.\n"
        "• generate_image — нарисовать изображение по описанию (промпт переводи на "
        "английский для качества). ⚠️ Если в запросе нет картинки и про изображение не "
        "просят — НЕ упоминай картинки и не проси их прислать, просто отвечай на вопрос.\n"
        "• analyze_image — посмотреть изображение (фото, скриншот, карточку) по пути к "
        "файлу или URL. Правила:\n"
        "    – Есть ссылка на картинку или просьба её описать → ВСЕГДА вызывай "
        "analyze_image, не отвечай «не могу посмотреть» (инструмент есть). Если ссылка "
        "была выше в диалоге — найди её в истории и передай.\n"
        "    – НИКОГДА не угадывай содержимое по имени файла или адресу. Сначала ПОСМОТРИ, "
        "потом описывай.\n"
        "    – 📊 СКРИНШОТ С ДАННЫМИ (таблица тарифов, прайс, отчёт, график, документ) — "
        "лучший источник истины. Переноси цифры как есть: не округляй, не путай "
        "колонки/строки, не дополняй из памяти; нечитаемое/обрезанное — честно отметь. "
        "Извлечённые со скрина числа называть МОЖНО. Для тарифов WB колонки «Склад WB "
        "(FBW), %», «Маркетплейс (FBS), %», «Витрина (DBS), %» — разные модели, не смешивай.\n"
        "• run_python — вычисления, анализ, проверка кода.\n\n"
        "🔎 КАЧЕСТВЕННЫЙ ПОИСК:\n"
        "• Запрос формулируй конкретно (ключевые слова, номера, год); для фактов добавляй "
        "«официальный»/«текст»/«2026». Не «оферта WB», а «оферта Wildberries пункт 9.7.2 текст».\n"
        "• Мало/мусорные результаты — переформулируй (синонимы, рус+англ) и попробуй снова. "
        "Один запрос — не приговор. Не говори «поищите сами» — это твоя работа.\n"
        "• Приоритет официальных источников (seller.wildberries.ru, .gov, nalog.ru, "
        "consultant.ru, документация) над блогами; при противоречии — верь официальному "
        "и честно укажи расхождение.\n"
        "• Фильтруй выдачу: используй ТОЛЬКО реально relevantное. Если пришёл мусор не по "
        "теме — не пересказывай его, скажи «релевантного не нашёл» и переформулируй или дай "
        "ответ из своих знаний (с пометкой). Отличай «упоминание» от «ответа»: если "
        "дословного ответа нет — так и скажи.\n"
        "• Нашёл источники — перечисли их в конце списком «Источники:».\n\n"
        "🛡️ ДОСТОВЕРНОСТЬ И ЧЕСТНОСТЬ (КРИТИЧНО):\n"
        "• 🔴 КРАСНАЯ ЛИНИЯ ПО ПРОВЕРЯЕМЫМ ДАННЫМ (для ЛЮБОЙ темы): конкретное проверяемое "
        "значение — процент (комиссии, тарифы, ставки налогов, штрафы, лимиты), цена, "
        "статистика, курс, дата события, чьё-то имя/должность, характеристика товара, "
        "цитата, номер дела/закона — называть КОНКРЕТНО можно ТОЛЬКО если в ЭТОМ ЖЕ ответе "
        "ты реально вызвал инструмент и значение есть в его результате. Нет прочитанного "
        "источника прямо сейчас — НЕ называй НИКАКОЕ точное значение, даже примерное, даже "
        "«обычно около» или «по моим данным». Пометка «требует проверки» НЕ легализует "
        "выдумку таких данных — значение просто не приводится. Выдуманный номер дела "
        "(типа «А41-183/2021») — опасная дезинформация, СТРОГО ЗАПРЕЩЕНО.\n"
        "• Что МОЖНО без инструмента: общеизвестные стабильные факты (столица Франции — "
        "Париж, вода кипит при 100 °C), общие принципы, объяснения «как устроено», общие "
        "формулировки права («товарные знаки в РФ защищены ГК РФ, часть 4» + «точную "
        "норму/практику уточни у юриста или в КонсультантПлюс»). Чего НЕЛЬЗЯ: свежие/"
        "изменчивые цифры (комиссии, цены, курсы, статистику года) без проверки.\n"
        "• Вместо выдуманного значения скажи, где его взять. Тарифы WB: «Точную "
        "комиссию из головы называть не буду — она зависит от категории и модели "
        "(FBW/FBS/DBS) и часто меняется. Смотри в WB Партнёры → Финансы → Тарифы → "
        "Комиссия (фильтр по категории) или в Калькуляторе тарифов. Пришли скриншот — "
        "разберу». Налоги — nalog.ru / личный кабинет. Прочее — предложи поискать сейчас "
        "(deep_research) или прислать источник/скриншот.\n"
        "• ⚠️ НЕЛЬЗЯ называть значение и рядом ставить источник, который ты НЕ открывал — "
        "это худший вид вранья (выглядит проверенным, но фейк). Источник указывай только "
        "если реально прочитал его инструментом в этом ответе. Пример галлюцинации: "
        "комиссия FBS в парфюмерии WB реально ~33–36% (по кабинету на 2026), а НЕ 17%.\n"
        "• ⚠️ ИЕРАРХИЯ ДОВЕРИЯ К ИСТОЧНИКАМ ЦИФР: личный кабинет/скриншот кабинета и "
        "официальный сайт > блоги/агентства/SEO-статьи. По тарифам, комиссиям, ставкам "
        "маркетплейсов блоги (mpagency, postavleno и т.п.) часто дают УСТАРЕВШИЕ или "
        "усреднённые цифры — а реальные тарифы в закрытом кабинете WB поиск прочитать НЕ "
        "может. Поэтому если нашёл цифру по тарифам только в блоге — НЕ подавай её как факт: "
        "пометь «по данным блогов ~X%, но это может быть устаревшим — точное только в твоём "
        "кабинете WB → Тарифы». Точную ставку всё равно веди смотреть в кабинет/по скриншоту.\n"
        "• ⚠️ СВЕРЯЙ С УЖЕ ИЗВЕСТНЫМ В ДИАЛОГЕ: если выше в этой беседе уже был более "
        "надёжный источник (скриншот кабинета, официальный сайт), а свежий поиск даёт ДРУГУЮ "
        "цифру — НЕ игнорируй конфликт. Опирайся на более надёжный источник и прямо отметь "
        "расхождение: «в блогах пишут ~10–20%, но по твоему скриншоту из кабинета реально "
        "33–36% — верь кабинету».\n"
        "• 🚫 НИКОГДА не утверждай, что что-то «сделал» (посмотрел сайт, нашёл, проверил, "
        "открыл), если не вызвал инструмент и не получил результат. Просят найти/посмотреть/"
        "проверить → сначала реально вызови инструмент, отвечай только по его результату. "
        "Не сработало/доступа нет/сайт закрыт → честно скажи, что именно не вышло, и "
        "предложи альтернативу (пусть пришлют ссылку/скриншот/файл; или дай общие знания, "
        "ЧЁТКО пометив, что это знания, а не результат проверки). Лучше честное «вот что "
        "удалось, а вот что нет», чем красивая фантазия.\n"
        "• 🧠 НЕ будь угодливым. Если пользователь сообщает факт/цифру, что противоречит "
        "твоим данным или звучит сомнительно — не поддакивай: вежливо отметь расхождение, "
        "по возможности перепроверь инструментом, уточни источник его цифры (из офиц. "
        "кабинета — верь; из сомнительного места — посоветуй перепроверить). Прав — спокойно "
        "согласись без лести. Не хватает данных — скажи «не могу подтвердить или "
        "опровергнуть». Не льсти без причины.\n\n"
        "📋 ФОРМАТ:\n"
        "• Не делай широкие таблицы (макс. 3-4 коротких колонки, только для коротких "
        "сравнений). Больше данных → формат КАРТОЧЕК: заголовок жирным, под ним список:\n"
        "  **1. Элегантный минимализм**\n"
        "  - Идея: премиальность и чистота\n"
        "  - Палитра: белый, золотой\n"
        "  Это читается в чате лучше громоздкой таблицы.\n"
        "• Длинный ответ разбивай на разделы с заголовками (#### Название), без воды и "
        "повторов. Структура research-ответа: вводная строка → разделы со списками → "
        "цифры с источником рядом → противоречия отметь честно → «Источники:» со ссылками "
        "и вывод/совет → уместный вопрос «Что дальше?».\n"
        "• ЯЗЫК: только грамотный русский, без вставок английских слов (используй русские "
        "эквиваленты). Проверяй текст перед ответом.\n\n"
        "💬 ПРИМЕРЫ:\n"
        "[Пользователь] привет → [Ты] Привет! 🙂 Чем помочь? (коротко, без инструментов)\n\n"
        "[Пользователь] 2+2? → [Ты] 4 🙂 (дай сам ответ, не «что ещё посчитать?»)\n\n"
        "[Пользователь] какая комиссия Wildberries в парфюмерии по FBS\n"
        "[Ты] (если НЕ открывал источник сейчас) Точный процент из головы называть не "
        "стану — он зависит от подкатегории (парфюмерная вода, наборы, селективный парфюм) "
        "и модели (FBW/FBS/DBS) и меняется. Надёжнее — кабинет: **WB Партнёры → Финансы → "
        "Тарифы → «Комиссия»**, фильтр по категории, колонка «Маркетплейс (FBS), %». Пришли "
        "скриншот — разберу. Или поискать актуальное в интернете прямо сейчас?\n"
        "— (НЕТ выдуманного «17%» и НЕТ фейковых ссылок)\n\n"
        "[Пользователь] (после поиска) какая комиссия WB в парфюмерии\n"
        "[Ты] (вызвал deep_research, реально прочитал) По данным <источник, что открыл>: "
        "FBS парфюмерной воды ~36%, FBW ~32,5% (2026). ⚠️ Часто меняется. Источники: "
        "<ссылка, что читал>.\n"
        "— (число есть ТОЛЬКО потому, что реально прочитан источник)\n\n"
        "[Пользователь] какие тренды в нейросетях в 2026\n"
        "[Ты] (deep_research → читаешь источники) Структурированный ответ с разделами и "
        "«Источники:» в конце.\n\n"
        "[Пользователь] помоги с карточкой товара\n"
        "[Ты] Конечно! Уточню: 1) что за товар? 2) что нужно — текст, дизайн, анализ? 🙂\n"
        "— (размытый запрос → сначала уточняешь)\n\n"
        "[Пользователь] посмотри топ-выдачу на Wildberries\n"
        "[Ты] (пробуешь web_search) Честно: зайти на WB и посмотреть выдачу напрямую не "
        "могу — сайт закрыт для авто-доступа. Но могу разобрать скриншоты или дать общие "
        "принципы (мои знания, не анализ карточек). Как удобнее?\n"
        "— (честно про ограничение + варианты, без выдумки)"
    )


    # Какие модели у каждого OpenAI-совместимого провайдера поддерживают tools.
    # Порядок = приоритет качества. gpt-oss грамотнее держит русский, чем llama-3.3.
    _FC_PROVIDERS = [
        ("groq", [
            "openai/gpt-oss-120b",       # грамотный русский, надёжный
            "openai/gpt-oss-20b",        # запасной, тоже gpt-oss
            "qwen/qwen3-32b",            # ещё запасной
            "llama-3.3-70b-versatile",   # llama в самом конце (бредит про картинки, слабее)
        ]),
        ("openrouter", [
            "openai/gpt-oss-120b:free",
            "openai/gpt-oss-20b:free",
            "qwen/qwen3-next-80b-a3b-instruct:free",
        ]),
        ("cerebras", ["llama3.3-70b"]),
        ("mistral", ["mistral-large-latest", "mistral-small-latest"]),
    ]

    # Алиасы для команды /model — короткое имя → (провайдер, model_id)
    MODEL_ALIASES = {
        "gpt-oss":   ("groq", "openai/gpt-oss-120b"),
        "llama":     ("groq", "llama-3.3-70b-versatile"),
        "fast":      ("groq", "llama-3.1-8b-instant"),
        "gemini":    ("google", None),
        "qwen":      ("openrouter", "qwen/qwen3-next-80b-a3b-instruct:free"),
    }

    def _run_via_openai_compat(self, msg: str, max_turns: int = 6) -> str | None:
        providers_list = self._FC_PROVIDERS
        # Если задана принудительная модель — используем только её
        if self.force_model and self.force_model in self.MODEL_ALIASES:
            pn, mid = self.MODEL_ALIASES[self.force_model]
            if pn != "google" and mid:
                providers_list = [(pn, [mid])]
            else:
                providers_list = []  # google обрабатывается отдельной веткой
        # Авто-режим: если на сессию уже закреплена рабочая модель — пробуем ЕЁ первой
        # (стабильность: одна модель ведёт весь диалог, без прыжков между сообщениями).
        elif self.session_model:
            sp, sm = self.session_model
            if sp != "google":
                rest = [(p, ms) for (p, ms) in self._FC_PROVIDERS]
                providers_list = [(sp, [sm])] + rest

        for prov_name, fc_models in providers_list:
            prov = self.pg.providers.get(prov_name)
            if not prov or not prov.ok:
                continue

            for model_id in fc_models:
                msgs = [
                    {"role": "system", "content": self.SYS_PROMPT},
                    {"role": "user", "content": msg},
                ]
                for _ in range(max_turns):
                    try:
                        kw = dict(model=model_id, messages=msgs, max_tokens=2048)
                        if self.tools:
                            kw["tools"] = self._tools_schema()
                            kw["tool_choice"] = "auto"

                        r = prov.c.chat.completions.create(**kw)
                        m = r.choices[0].message

                        if m.tool_calls:
                            tcs = [{"id": tc.id, "type": "function",
                                    "function": {"name": tc.function.name,
                                                 "arguments": tc.function.arguments}}
                                   for tc in m.tool_calls]
                            msgs.append({"role": "assistant", "content": m.content or "", "tool_calls": tcs})
                            for tc in m.tool_calls:
                                try:
                                    args = json.loads(tc.function.arguments or "{}")
                                except Exception:
                                    args = {}
                                result = self._execute_tool(tc.function.name, args)
                                if self.debug:
                                    print(f"   [{prov_name} tool] {tc.function.name}({args}) -> {result[:120]}")
                                msgs.append({"role": "tool", "tool_call_id": tc.id, "content": result})
                        else:
                            self.last_model = f"{prov_name}/{model_id}"
                            # Закрепляем рабочую модель на сессию (только в авто-режиме)
                            if not self.force_model:
                                self.session_model = (prov_name, model_id)
                            return m.content or "[no response]"
                    except Exception as e:
                        if self.debug:
                            print(f"   [{prov_name} {model_id} error] {type(e).__name__}: {str(e)[:120]}")
                        break  # эта модель не сработала — следующая
        return None

    def run(self, msg: str, max_turns: int = 8) -> str:
        import re as _re
        IMG_RE = r'https?://\S+\.(?:jpg|jpeg|png|webp|gif|bmp)(?:\?\S*)?'

        # Находим все ссылки на картинки во всём сообщении (включая историю диалога).
        img_urls = _re.findall(IMG_RE, msg, _re.I)
        # Признак, что пользователь говорит про изображение
        wants_image = bool(_re.search(r'(?i)(картинк|изображени|фото|снимок|скрин|опиши.*(её|ее|это)|на ней|на нём|на нем)', msg))

        if img_urls:
            # Берём ПОСЛЕДНЮЮ упомянутую ссылку на картинку (самую свежую в диалоге)
            last_img = img_urls[-1]
            msg = msg + (f"\n\n[Система: в диалоге есть изображение: {last_img}\n"
                         f"Если пользователь просит описать/проанализировать картинку — "
                         f"ОБЯЗАТЕЛЬНО вызови инструмент analyze_image с url='{last_img}'. "
                         f"НЕ придумывай содержимое по названию файла — посмотри реально через инструмент.]")
        elif wants_image:
            msg = msg + ("\n\n[Система: пользователь спрашивает про изображение, но прямой ссылки нет. "
                         "Посмотри выше в диалоге — возможно, ссылка была в предыдущих сообщениях. "
                         "Если нашёл — вызови analyze_image. Если ссылки нигде нет — попроси прислать её.]")


        # Для картинок Gemini надёжнее (видит изображения нативно) — пробуем его ПЕРВЫМ,
        # даже если выбрана другая модель.
        if (img_urls or wants_image) and self.force_model != "gemini":
            gprov = self.pg.providers.get("google")
            if gprov and gprov.ok:
                result = self._run_via_google(msg, max_turns)
                if result is not None:
                    return result

        # Если принудительно выбран Gemini — сначала пробуем его
        if self.force_model == "gemini":
            result = self._run_via_google(msg, max_turns)
            if result is not None:
                return result

        # 1. OpenAI-совместимые провайдеры: Groq → OpenRouter → Cerebras → Mistral
        result = self._run_via_openai_compat(msg, max_turns)
        if result is not None:
            return result

        # 2. Google Gemini (нативный function calling) — как резерв
        result = self._run_via_google(msg, max_turns)
        if result is not None:
            return result

        # 3. Fallback: ответ без инструментов через роутинг
        self.last_model = "fallback/route"
        return self.pg.route(msg)



# ===== Стандартные инструменты =====

def default_tools(tavily_key: str = "", workdir: str = "agent_files", allow_exec: bool = True) -> list[Tool]:
    import subprocess, sys, urllib.parse, urllib.request

    os.makedirs(workdir, exist_ok=True)

    def _safe_path(path: str) -> str:
        """Не даём вырваться за пределы рабочей папки агента."""
        full = os.path.abspath(os.path.join(workdir, path))
        base = os.path.abspath(workdir)
        if not full.startswith(base):
            raise ValueError("путь вне рабочей папки запрещён")
        return full

    def search(query: str) -> str:
        # 1) Tavily (если есть ключ) — самый качественный поиск для AI
        if tavily_key:
            try:
                import requests
                r = requests.post("https://api.tavily.com/search",
                    json={"api_key": tavily_key, "query": query,
                          "search_depth": "advanced", "max_results": 6,
                          "include_answer": True}, timeout=20)
                data = r.json()
                parts = []
                if data.get("answer"):
                    parts.append(f"Сводка: {data['answer']}")
                for it in data.get("results", [])[:6]:
                    parts.append(f"- {it.get('title','')} ({it.get('url','')}): "
                                 f"{it.get('content','')[:300]}")
                if parts:
                    return "\n".join(parts)
            except Exception:
                pass

        out = []
        # 2) DuckDuckGo Instant Answer — хорош для определений/фактов
        try:
            url = "https://api.duckduckgo.com/?q=" + urllib.parse.quote(query) + "&format=json&no_html=1"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            data = json.loads(urllib.request.urlopen(req, timeout=12).read().decode("utf-8", "ignore"))
            if data.get("AbstractText"):
                out.append(f"- {data.get('Heading','')}: {data['AbstractText']}")
            for t in data.get("RelatedTopics", [])[:4]:
                if isinstance(t, dict) and t.get("Text"):
                    out.append(f"- {t['Text']}")
        except Exception:
            pass

        # 3) Wikipedia (ru) — надёжный бесплатный источник без ключа
        if len(out) < 3:
            try:
                import re as _re
                wurl = "https://ru.wikipedia.org/w/api.php?" + urllib.parse.urlencode({
                    "action":"query","list":"search","srsearch":query,
                    "format":"json","srlimit":5,
                })
                wreq = urllib.request.Request(wurl, headers={"User-Agent":"Polygraph/1.0"})
                wdata = json.loads(urllib.request.urlopen(wreq, timeout=12).read().decode("utf-8","ignore"))
                for it in wdata.get("query",{}).get("search",[]):
                    snip = _re.sub(r"<[^>]+>", "", it.get("snippet",""))
                    out.append(f"- {it.get('title','')}: {snip}")
            except Exception:
                pass

        if out:
            # убираем дубли, ограничиваем
            seen, uniq = set(), []
            for line in out:
                if line not in seen:
                    seen.add(line); uniq.append(line)
            return "\n".join(uniq[:8])
        return (f"[поиск '{query}'] — ничего не найдено. Совет: используй fetch_url "
                f"с конкретной ссылкой, либо задай TAVILY_API_KEY для лучшего поиска.")

    def calc(expression: str) -> str:
        try:
            return str(eval(expression, {"__builtins__":{}}, {
                "abs":abs,"round":round,"min":min,"max":max,"pow":pow,"sum":sum,"len":len}))
        except Exception as e:
            return f"ошибка: {e}"

    def read_file(path: str) -> str:
        try:
            full = _safe_path(path)
            if not os.path.exists(full):
                return f"[файл не найден: {path}]"
            # PDF — извлекаем текст (если установлен pypdf)
            if full.lower().endswith(".pdf"):
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(full)
                    text = "\n".join((pg.extract_text() or "") for pg in reader.pages[:30])
                    return text[:8000] if text.strip() else "[PDF без извлекаемого текста (возможно, скан — попробуй analyze_image для страниц)]"
                except ImportError:
                    return "[для чтения PDF установи библиотеку: pip install pypdf]"
                except Exception as e:
                    return f"[ошибка чтения PDF: {e}]"
            # Excel (.xlsx / .xls) — извлекаем таблицу как текст
            if full.lower().endswith((".xlsx", ".xls", ".xlsm")):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(full, read_only=True, data_only=True)
                    out = []
                    for ws in wb.worksheets[:5]:
                        out.append(f"=== Лист: {ws.title} ===")
                        rows = 0
                        for row in ws.iter_rows(values_only=True):
                            cells = ["" if c is None else str(c) for c in row]
                            if any(cells):
                                out.append(" | ".join(cells))
                                rows += 1
                            if rows >= 200:
                                out.append("... (показаны первые 200 строк)")
                                break
                    txt = "\n".join(out)
                    return txt[:8000] if txt.strip() else "[Excel-файл пуст]"
                except ImportError:
                    return "[для чтения Excel установи: pip install openpyxl]"
                except Exception as e:
                    return f"[ошибка чтения Excel: {e}]"
            # CSV — читаем как таблицу
            if full.lower().endswith(".csv"):
                try:
                    import csv as _csv
                    out = []
                    for enc in ("utf-8-sig", "cp1251", "utf-8"):
                        try:
                            with open(full, "r", encoding=enc, newline="") as f:
                                sample = f.read(2048); f.seek(0)
                                delim = ";" if sample.count(";") > sample.count(",") else ","
                                reader = _csv.reader(f, delimiter=delim)
                                for i, row in enumerate(reader):
                                    out.append(" | ".join(row))
                                    if i >= 200:
                                        out.append("... (первые 200 строк)")
                                        break
                            break
                        except UnicodeDecodeError:
                            out = []; continue
                    txt = "\n".join(out)
                    return txt[:8000] if txt.strip() else "[CSV пуст]"
                except Exception as e:
                    return f"[ошибка чтения CSV: {e}]"
            # Обычный текстовый файл
            with open(full, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            return content[:8000] if content else "[файл пуст]"
        except Exception as e:
            return f"[ошибка чтения: {e}]"

    def write_file(path: str, content: str) -> str:
        try:
            p = _safe_path(path)
            os.makedirs(os.path.dirname(p) or ".", exist_ok=True)
            with open(p, "w", encoding="utf-8") as f:
                f.write(content)
            return f"[записано {len(content)} символов в {path}]"
        except Exception as e:
            return f"[ошибка записи: {e}]"

    def list_files(path: str = ".") -> str:
        try:
            base = _safe_path(path)
            items = []
            for name in sorted(os.listdir(base)):
                full = os.path.join(base, name)
                size = os.path.getsize(full) if os.path.isfile(full) else "-"
                items.append(f"{'DIR ' if os.path.isdir(full) else 'FILE'} {name} ({size})")
            return "\n".join(items) or "[папка пуста]"
        except Exception as e:
            return f"[ошибка: {e}]"

    def run_python(code: str) -> str:
        """Выполнить Python-код в подпроцессе (внутри рабочей папки)."""
        try:
            r = subprocess.run([sys.executable, "-c", code], capture_output=True,
                               text=True, timeout=20, cwd=workdir)
            out = (r.stdout or "")[:4000]
            err = (r.stderr or "")[:2000]
            res = ""
            if out: res += f"STDOUT:\n{out}"
            if err: res += f"\nSTDERR:\n{err}"
            return res.strip() or "[код выполнен, вывода нет]"
        except subprocess.TimeoutExpired:
            return "[ошибка: превышено время выполнения 20с]"
        except Exception as e:
            return f"[ошибка выполнения: {e}]"

    def fetch_url(url: str) -> str:
        """Скачать веб-страницу и вернуть её текст (HTML → чистый текст)."""
        import re, html as html_mod
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        # Кодируем не-ASCII символы в URL (кириллица в адресе и т.п.)
        try:
            url = urllib.parse.quote(url, safe=":/?#[]@!$&'()*+,;=%~")
        except Exception:
            pass
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/120.0 Safari/537.36",
                "Accept-Language": "ru,en;q=0.9",
            })
            raw = urllib.request.urlopen(req, timeout=20).read()
            # Если это PDF (по сигнатуре или расширению) — извлекаем текст через pypdf
            if raw[:5] == b"%PDF-" or url.lower().split("?")[0].endswith(".pdf"):
                try:
                    from pypdf import PdfReader
                    import io as _io
                    reader = PdfReader(_io.BytesIO(raw))
                    pdftext = "\n".join((pg.extract_text() or "") for pg in reader.pages[:40])
                    if pdftext.strip():
                        return pdftext[:9000]
                    return f"[PDF по ссылке {url} открыт, но текст не извлёкся (возможно, скан-картинки)]"
                except ImportError:
                    return "[для чтения PDF установи: pip install pypdf]"
                except Exception as e:
                    return f"[ошибка чтения PDF: {e}]"
            try:
                text = raw.decode("utf-8")
            except UnicodeDecodeError:
                text = raw.decode("cp1251", "ignore")
        except Exception as e:
            return f"[не удалось открыть {url}: {e}]"

        # Удаляем скрипты/стили/теги, превращаем в читаемый текст
        text = re.sub(r"(?is)<(script|style|noscript|svg|head|nav|header|footer|aside|form|button|select).*?</\1>", " ", text)
        # Вырезаем типовые блоки навигации/меню по id/class
        text = re.sub(r'(?is)<(div|ul|section)[^>]*(id|class)="[^"]*(nav|menu|sidebar|footer|header|banner|cookie|lang)[^"]*"[^>]*>.*?</\1>', " ", text)
        text = re.sub(r"(?is)<br\s*/?>", "\n", text)
        text = re.sub(r"(?is)</(p|div|li|tr|h[1-6])>", "\n", text)
        text = re.sub(r"(?is)<[^>]+>", " ", text)            # остальные теги
        text = html_mod.unescape(text)                        # &amp; → &
        # Подчищаем вики-разметку, если попалась: [[ссылка|текст]] → текст, <ref>…</ref> → ''
        text = re.sub(r"(?is)<ref[^>]*>.*?</ref>", "", text)
        text = re.sub(r"(?is)<ref[^>]*/>", "", text)
        text = re.sub(r"\[\[(?:[^\]|]*\|)?([^\]]+)\]\]", r"\1", text)
        text = re.sub(r"\{\{[^}]*\}\}", "", text)
        text = re.sub(r"<!--.*?-->", "", text, flags=re.S)
        text = re.sub(r"[ \t]+", " ", text)

        # Оставляем содержательные строки: длинные ИЛИ с точкой/двоеточием.
        # Это отсекает списки коротких ссылок (меню, список языков и т.п.).
        good_lines = []
        for ln in text.split("\n"):
            s = ln.strip()
            if not s:
                continue
            if len(s) >= 40 or s.endswith((".", ":", "!", "?")) or s.count(" ") >= 6:
                good_lines.append(s)
        text = "\n".join(good_lines)
        text = re.sub(r"\n\s*\n+", "\n\n", text).strip()
        if not text:
            return f"[страница {url} открыта, но текст не извлёкся (возможно, JS-сайт)]"
        return text[:6000]  # ограничиваем, чтобы влезло в контекст модели

    def analyze_image(image: str, question: str = "Опиши подробно, что на этом изображении.") -> str:
        """Проанализировать изображение (путь к файлу в рабочей папке или URL).
        Использует Gemini (vision) или vision-модель OpenRouter."""
        import base64, urllib.request as _u

        # Получаем байты картинки + mime
        img_bytes, mime = None, "image/jpeg"
        try:
            if image.startswith(("http://", "https://")):
                req = _u.Request(image, headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    "Accept": "image/avif,image/webp,image/png,image/jpeg,image/*,*/*;q=0.8",
                    "Referer": "https://www.google.com/",
                })
                img_bytes = _u.urlopen(req, timeout=20).read()
                if image.lower().endswith(".png"): mime = "image/png"
                elif image.lower().endswith((".webp",)): mime = "image/webp"
            else:
                p = _safe_path(image)
                if not os.path.exists(p):
                    return f"[изображение не найдено: {image}]"
                with open(p, "rb") as f:
                    img_bytes = f.read()
                if p.lower().endswith(".png"): mime = "image/png"
                elif p.lower().endswith(".webp"): mime = "image/webp"
        except Exception as e:
            return (f"ОШИБКА ЗАГРУЗКИ ИЗОБРАЖЕНИЯ ({e}). "
                    f"Сообщи пользователю ЧЕСТНО: картинку по этой ссылке скачать не удалось "
                    f"(возможно, сайт блокирует доступ). НЕ ПРИДУМЫВАЙ, что на ней изображено — "
                    f"ты её НЕ видел. Попроси прислать другую ссылку или загрузить файл.")

        if not img_bytes:
            return "[пустое изображение]"

        # Определяем MIME по сигнатуре файла (надёжнее, чем по расширению URL)
        if img_bytes[:3] == b"\xff\xd8\xff":
            mime = "image/jpeg"
        elif img_bytes[:8] == b"\x89PNG\r\n\x1a\n":
            mime = "image/png"
        elif img_bytes[:4] == b"RIFF" and img_bytes[8:12] == b"WEBP":
            mime = "image/webp"
        elif img_bytes[:6] in (b"GIF87a", b"GIF89a"):
            mime = "image/gif"

        # 1) Пробуем Gemini (нативный vision). Несколько моделей — у каждой свой лимит.
        gkey = os.environ.get("GEMINI_API_KEY", "")
        if gkey:
            try:
                from google import genai
                from google.genai import types as gt
                client = genai.Client(api_key=gkey)
                for gmodel in ("gemini-2.5-flash", "gemini-2.0-flash", "gemini-2.5-pro"):
                    try:
                        resp = client.models.generate_content(
                            model=gmodel,
                            contents=[
                                gt.Part.from_bytes(data=img_bytes, mime_type=mime),
                                question,
                            ],
                        )
                        if resp and getattr(resp, "text", None):
                            return resp.text
                    except Exception:
                        continue
            except Exception:
                pass

        # 2) OpenRouter vision (qwen2.5-vl free)
        okey = os.environ.get("OPENROUTER_API_KEY", "")
        if okey:
            try:
                from openai import OpenAI
                client = OpenAI(api_key=okey, base_url="https://openrouter.ai/api/v1")
                b64 = base64.b64encode(img_bytes).decode()
                for vmodel in ("google/gemma-4-31b-it:free",
                               "google/gemma-4-26b-a4b-it:free",
                               "nvidia/nemotron-nano-12b-v2-vl:free",
                               "moonshotai/kimi-k2.6:free"):
                    try:
                        r = client.chat.completions.create(
                            model=vmodel,
                            messages=[{"role": "user", "content": [
                                {"type": "text", "text": question},
                                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                            ]}],
                            max_tokens=1024, timeout=40)
                        c = r.choices[0].message.content
                        if c:
                            return c
                    except Exception:
                        continue
            except Exception:
                pass

        return ("Не получилось проанализировать изображение прямо сейчас — vision-модели "
                "(Gemini / OpenRouter) недоступны: вероятно, исчерпан дневной лимит или "
                "провайдер занят. Сообщи это пользователю ЧЕСТНО и предложи попробовать "
                "чуть позже. НЕ придумывай содержимое картинки.")

    def generate_image(prompt: str) -> str:
        """Генерация изображения. Gemini (nano-banana) → Cloudflare FLUX как запаска.
        Возвращает маркер IMAGE: для показа в чате."""
        import time as _t, base64 as _b64
        gendir = os.path.join(workdir, "generated")
        os.makedirs(gendir, exist_ok=True)
        fname = f"img_{int(_t.time())}.png"
        fpath = os.path.join(gendir, fname)
        rel = os.path.join("generated", fname).replace("\\", "/")
        errors = []

        # 1) Gemini (nano-banana) — ~500/день бесплатно, нужен VPN
        gkey = os.environ.get("GEMINI_API_KEY", "")
        if gkey:
            try:
                from google import genai
                from google.genai import types as gt
                client = genai.Client(api_key=gkey)
                try:
                    resp = client.models.generate_content(
                        model="gemini-2.5-flash-image", contents=prompt,
                        config=gt.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                    )
                    parts = resp.candidates[0].content.parts if resp.candidates else []
                    for part in parts:
                        if getattr(part, "inline_data", None) and part.inline_data.data:
                            with open(fpath, "wb") as f:
                                f.write(part.inline_data.data)
                            return f"IMAGE:{rel}|Готово! Нарисовал по запросу: «{prompt}»."
                    errors.append("Gemini: ответ без картинки")
                except Exception as e:
                    m = str(e)
                    if "429" in m or "RESOURCE_EXHAUSTED" in m: errors.append("Gemini: лимит (429)")
                    elif "403" in m or "PERMISSION" in m or "Forbidden" in m: errors.append("Gemini: доступ (403, VPN/биллинг)")
                    else: errors.append(f"Gemini: {m[:60]}")
            except Exception as e:
                errors.append(f"Gemini-клиент: {str(e)[:60]}")

        # 2) Cloudflare Workers AI (FLUX) — бесплатно, не зависит от Gemini
        cf_acc = os.environ.get("CLOUDFLARE_ACCOUNT_ID", "")
        cf_tok = os.environ.get("CLOUDFLARE_API_TOKEN", "")
        if cf_acc and cf_tok:
            try:
                import urllib.request as _u, json as _j
                url = (f"https://api.cloudflare.com/client/v4/accounts/{cf_acc}"
                       f"/ai/run/@cf/black-forest-labs/flux-1-schnell")
                body = _j.dumps({"prompt": prompt[:2000]}).encode()
                req = _u.Request(url, data=body, headers={
                    "Authorization": f"Bearer {cf_tok}", "Content-Type": "application/json"})
                resp = _u.urlopen(req, timeout=60).read()
                data = _j.loads(resp)
                img_b64 = data.get("result", {}).get("image")
                if img_b64:
                    with open(fpath, "wb") as f:
                        f.write(_b64.b64decode(img_b64))
                    return f"IMAGE:{rel}|Готово! Нарисовал по запросу: «{prompt}» (FLUX)."
                errors.append("Cloudflare: пустой ответ")
            except Exception as e:
                errors.append(f"Cloudflare: {str(e)[:60]}")

        diag = " | ".join(errors) if errors else "нет доступного генератора"
        hint = ""
        if not gkey and not cf_acc:
            hint = " Добавь GEMINI_API_KEY (с VPN) или CLOUDFLARE_ACCOUNT_ID+CLOUDFLARE_API_TOKEN в .env."
        elif "429" in diag:
            hint = " Лимит Gemini исчерпан — подожди сброса (раз в сутки) или добавь Cloudflare-ключ как запаску."
        return ("Не удалось сгенерировать изображение. Причина: " + diag + "." + hint +
                " (Я не выдумываю картинку — честно сообщаю результат.)")

    def deep_research(query: str) -> str:
        """Глубокое исследование (как ресёрчер): несколько поисковых запросов +
        чтение топ-страниц с приоритетом официальных источников. Возвращает
        собранный материал со ссылками — модель делает честный вывод."""
        import re as _re
        collected = []
        sources = []
        seen_urls = set()

        # 1) Делаем НЕСКОЛЬКО запросов: основной + вариации (как человек уточняет).
        queries = [query]
        # добавим вариацию со словами, повышающими шанс найти точные данные
        if not _re.search(r'(?i)(официальн|точн|текст|документ)', query):
            queries.append(query + " официальный источник 2026")

        for q in queries[:2]:
            res = search(q)
            if res and "ничего не найдено" not in res:
                collected.append(f"=== Поиск: «{q}» ===\n{res}")

        # 2) Если есть Tavily — собираем URL'ы из всех запросов и читаем топ-страницы,
        #    отдавая ПРИОРИТЕТ официальным доменам (gov, edu, офиц. сайты компаний).
        if tavily_key:
            try:
                import requests
                all_urls = []
                for q in queries[:2]:
                    try:
                        r = requests.post("https://api.tavily.com/search",
                            json={"api_key": tavily_key, "query": q,
                                  "search_depth": "advanced", "max_results": 5,
                                  "include_answer": False}, timeout=20)
                        for it in r.json().get("results", []):
                            u = it.get("url")
                            if u and u not in seen_urls:
                                all_urls.append(u); seen_urls.add(u)
                    except Exception:
                        continue
                # Приоритет: официальные источники выше
                def _rank(u):
                    ul = u.lower()
                    if any(d in ul for d in (".gov", ".gov.ru", "seller.wildberries", "ozon.ru",
                                             "official", "minfin", "nalog.ru", "consultant.ru",
                                             "garant.ru", "docs.", "developer")):
                        return 0  # офиц. — первыми
                    if any(d in ul for d in (".edu", "wikipedia", "habr.com")):
                        return 1
                    return 2
                all_urls.sort(key=_rank)
                for u in all_urls[:3]:
                    page = fetch_url(u)
                    if page and not page.startswith("["):
                        collected.append(f"\n=== Содержимое {u} ===\n{page[:2500]}")
                        sources.append(u)
            except Exception as e:
                collected.append(f"[deep: чтение страниц пропущено: {e}]")
        else:
            collected.append("[Совет: задай TAVILY_API_KEY для чтения полных страниц.]")

        result = "\n".join(collected)
        if sources:
            result += "\n\n=== ИСТОЧНИКИ (укажи их в ответе) ===\n" + \
                      "\n".join(f"[{i+1}] {s}" for i, s in enumerate(sources))
        result += ("\n\n[ПАМЯТКА: используй ТОЛЬКО факты из текста выше. Если точного "
                   "ответа/числа здесь НЕТ — честно скажи об этом и НЕ выдумывай. "
                   "Отличай 'упоминание темы' от 'дословного ответа на вопрос'. "
                   "Предложи, как получить точные данные (офиц. сайт/ссылка/скриншот).]")
        return result[:9000]

    tools = [
        Tool("generate_image", "Сгенерировать (нарисовать) изображение по текстовому описанию. Используй, когда просят нарисовать, создать картинку, логотип, иллюстрацию, баннер, концепт и т.п. ВАЖНО: prompt передавай на АНГЛИЙСКОМ языке (переведи запрос пользователя) — так качество намного лучше. Будь подробным в описании (стиль, цвета, детали).",
            {"type":"object","properties":{"prompt":{"type":"string","description":"Подробное описание картинки на английском"}},"required":["prompt"]},
            generate_image),
        Tool("deep_research", "ГЛУБОКОЕ исследование темы: делает поиск и читает несколько страниц целиком, возвращает материал со ссылками. Используй для сложных вопросов, требующих анализа из нескольких источников ('найди и сделай вывод', обзоры, сравнения, актуальные темы).",
            {"type":"object","properties":{"query":{"type":"string","description":"Тема/вопрос для исследования"}},"required":["query"]},
            deep_research),
        Tool("web_search", "Поиск актуальной информации в интернете",
            {"type":"object","properties":{"query":{"type":"string","description":"Поисковый запрос"}},"required":["query"]},
            search),
        Tool("calculator", "Математические вычисления",
            {"type":"object","properties":{"expression":{"type":"string","description":"Выражение: 2+2*10"}},"required":["expression"]},
            calc),
        Tool("datetime_now", "Текущие дата и время",
            {"type":"object","properties":{}},
            lambda: time.strftime("%Y-%m-%d %H:%M:%S")),
        Tool("read_file", "Прочитать содержимое файла из рабочей папки агента. Поддерживает текст, PDF, Excel (.xlsx/.xls) и CSV — таблицы извлекаются как текст для анализа.",
            {"type":"object","properties":{"path":{"type":"string","description":"Имя файла, напр. notes.txt"}},"required":["path"]},
            read_file),
        Tool("write_file", "Создать/перезаписать файл в рабочей папке агента",
            {"type":"object","properties":{
                "path":{"type":"string","description":"Имя файла"},
                "content":{"type":"string","description":"Содержимое файла"}},"required":["path","content"]},
            write_file),
        Tool("list_files", "Показать список файлов в рабочей папке агента",
            {"type":"object","properties":{"path":{"type":"string","description":"Подпапка (по умолчанию текущая)"}}},
            list_files),
        Tool("analyze_image", "Проанализировать изображение (фото, скриншот, карточку товара, инфографику). Принимает путь к файлу в рабочей папке агента ИЛИ URL картинки. Можно задать вопрос об изображении. Используй, когда нужно 'посмотреть' картинку, описать её, прочитать текст с неё, проанализировать дизайн.",
            {"type":"object","properties":{
                "image":{"type":"string","description":"Путь к файлу (напр. photo.jpg) или URL изображения"},
                "question":{"type":"string","description":"Что узнать об изображении (необязательно)"}},"required":["image"]},
            analyze_image),
        Tool("fetch_url", "Открыть веб-страницу по URL и получить её текстовое содержимое для анализа. Используй, когда нужно прочитать конкретную страницу или статью.",
            {"type":"object","properties":{"url":{"type":"string","description":"Ссылка на страницу, напр. https://example.com/page"}},"required":["url"]},
            fetch_url),
    ]

    if allow_exec:
        tools.append(Tool("run_python", "Выполнить Python-код и вернуть его вывод. Используй для вычислений, анализа данных, проверки кода.",
            {"type":"object","properties":{"code":{"type":"string","description":"Python-код для выполнения"}},"required":["code"]},
            run_python))

    return tools

